# -*- coding: utf-8 -*-
"""
Genera data/processed/ a partir de data/raw/BBDD_Actores_reordenada_v2.xlsx.

Salidas:
  1. data/processed/csv/            un CSV tidy (UTF-8) por pestaña
  2. data/processed/BBDD_Actores_procesada.xlsx
       - copia del libro original con dos correcciones de fecha documentadas
       - pestaña RESUMEN agregada al inicio

Correcciones aplicadas (ver data/data_dictionary.md, sección "Inconsistencias"):
  - "Foros - Índice de Eventos": dos celdas de fecha guardadas como serial de
    Excel interpretado en 1905 (efecto de escribir solo el año en una celda con
    formato fecha). "4° Lithium Latin America Congress" -> 2023;
    "LatAm & Argentina Critical Minerals Summit 2025" -> 2025.

Uso:  python scripts/build_processed.py   (desde la raíz del repo)
"""

import csv
import datetime
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw" / "BBDD_Actores_reordenada_v2.xlsx"
OUT_XLSX = ROOT / "data" / "processed" / "BBDD_Actores_procesada.xlsx"
OUT_CSV = ROOT / "data" / "processed" / "csv"

THEMATIC = [
    "Empresas Mineras",
    "Instituciones Estatales",
    "Comunidades Territoriales",
    "ONGs y ambiente",
    "Academia e Investigacion",
    "Organismos Internacionales",
    "Actores Financieros",
    "Servicios Prof. y Consultoras",
    "Sindicatos y Gremios",
]

DATE_FIXES = {  # evento -> año real (celda original: serial de Excel en 1905)
    "4° Lithium Latin America Congress": "2023",
    "LatAm & Argentina Critical Minerals Summit 2025": "2025",
}


def slug(name):
    s = name.lower()
    for a, b in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ñ", "n")]:
        s = s.replace(a, b)
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s


def fmt(value):
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    return str(value).strip()


def sheet_rows(wb, name):
    rows = [list(r) for r in wb[name].iter_rows(values_only=True)]
    width = max((len(r) for r in rows), default=0)
    return [r + [None] * (width - len(r)) for r in rows]


def is_empty(row):
    return all(fmt(c) == "" for c in row)


def write_csv(path, header, rows):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print(f"  {path.relative_to(ROOT)}  ({len(rows)} filas)")


def is_structural(first_cell, rest):
    """Fila decorativa: banner de sección (►), título en mayúsculas o
    subencabezado '└─ EN MAYÚSCULAS' sin datos en las demás columnas."""
    if first_cell.startswith("►"):
        return True
    if any(v != "" for v in rest):
        return False
    label = first_cell.lstrip("└─ ").strip()
    letters = [c for c in label if c.isalpha()]
    return bool(letters) and label == label.upper()


def tidy_sectioned(rows):
    """Pestañas con banners ► y fila de encabezado 'Nombre'."""
    header, data, seccion = None, [], ""
    for row in rows:
        c0 = fmt(row[0])
        rest = [fmt(c) for c in row[1:]]
        if header is None:
            if c0 == "Nombre":
                header = [fmt(c) for c in row]
                while header and header[-1] == "":
                    header.pop()
            continue
        if is_empty(row):
            continue
        if c0.startswith("►"):
            seccion = c0.lstrip("► ").strip()
            continue
        if is_structural(c0, rest):
            continue
        if c0 == "" and all(v == "" for v in rest[: len(header)]):
            continue
        name = c0.lstrip("└─ ").strip() if c0 else ""
        record = [name] + [fmt(c) for c in row[1: len(header)]]
        data.append([seccion] + record)
    return ["seccion"] + header, data


def tidy_flat(rows, fixes=None):
    """Pestañas de foros: encabezado en la fila 1, datos directos."""
    header = [fmt(c) for c in rows[0]]
    while header and header[-1] == "":
        header.pop()
    data = []
    for row in rows[1:]:
        if is_empty(row):
            continue
        rec = [fmt(c) for c in row[: len(header)]]
        if fixes:
            rec = fixes(rec)
        data.append(rec)
    return header, data


def tidy_matrix(rows):
    """'Multiafiliación -': bloques de 3 filas por actor
    (organizaciones / cargos / períodos) bajo grupos sectoriales definidos
    por la primera fila. Se exporta en formato largo."""
    nonempty = [r for r in rows if not is_empty(r)]
    sector_row = nonempty[0]
    labels = [(i, fmt(c)) for i, c in enumerate(sector_row) if fmt(c)]
    ranges = []
    for k, (i, lab) in enumerate(labels):
        end = labels[k + 1][0] if k + 1 < len(labels) else len(sector_row)
        ranges.append((i, end, lab))

    def sector_of(col):
        for start, end, lab in ranges:
            if start <= col < end:
                return lab
        return ""

    data, i = [], 1
    while i < len(nonempty):
        row = nonempty[i]
        actor = fmt(row[0])
        if not actor:
            i += 1
            continue
        cargos = nonempty[i + 1] if i + 1 < len(nonempty) and fmt(nonempty[i + 1][0]) == "" else []
        periodos = nonempty[i + 2] if i + 2 < len(nonempty) and cargos != [] and fmt(nonempty[i + 2][0]) == "" else []
        for col in range(1, len(row)):
            org = fmt(row[col])
            if not org:
                continue
            cargo = fmt(cargos[col]) if cargos and col < len(cargos) else ""
            periodo = fmt(periodos[col]) if periodos and col < len(periodos) else ""
            data.append([actor, sector_of(col), org, cargo.strip("()"), periodo])
        i += 1 + (1 if cargos != [] else 0) + (1 if periodos != [] else 0)
    return ["actor", "sector", "organizacion", "cargo", "periodo"], data


def build_csvs(wb):
    OUT_CSV.mkdir(parents=True, exist_ok=True)
    counts = {}
    for name in THEMATIC + ["Cruces Multiafiliacion"]:
        header, data = tidy_sectioned(sheet_rows(wb, name))
        write_csv(OUT_CSV / f"{slug(name)}.csv", header, data)
        counts[name] = len(data)

    def fix_indice(rec):
        if rec[0] in DATE_FIXES:
            rec[2] = DATE_FIXES[rec[0]]
        return rec

    for name, fixes in [
        ("Foros- Base Completa", None),
        ("Foros - Índice de Eventos", fix_indice),
        ("Foros - Nodos Centrales", None),
    ]:
        header, data = tidy_flat(sheet_rows(wb, name), fixes)
        write_csv(OUT_CSV / f"{slug(name)}.csv", header, data)
        counts[name] = len(data)

    header, data = tidy_matrix(sheet_rows(wb, "Multiafiliación -"))
    write_csv(OUT_CSV / "multiafiliacion_matriz_larga.csv", header, data)
    counts["Multiafiliación -"] = len(data)
    return counts


RESUMEN_SPEC = [
    # pestaña, tipo, descripción breve, aporte de red
    ("Academia e Investigacion", "Temática", "Universidades, centros, investigadores y think tanks", "Atributos de nodo"),
    ("Empresas Mineras", "Temática", "Operadoras, ejecutivos, cadena de valor y tecnología DLE", "Atributos de nodo"),
    ("Instituciones Estatales", "Temática", "Ministerios, agencias, reguladores y funcionarios", "Atributos de nodo"),
    ("ONGs y ambiente", "Temática", "ONGs y organizaciones ambientales", "Atributos de nodo"),
    ("Comunidades Territoriales", "Temática", "Comunidades y dirigencias indígenas", "Atributos de nodo"),
    ("Organismos Internacionales", "Temática", "Multilaterales, cooperación y banca de desarrollo", "Atributos de nodo"),
    ("Servicios Prof. y Consultoras", "Temática", "Bufetes, consultoras y servicios profesionales", "Atributos de nodo"),
    ("Actores Financieros", "Temática", "Bancos, fondos y asesores financieros", "Atributos de nodo"),
    ("Sindicatos y Gremios", "Temática", "Gremios empresariales y sindicatos", "Atributos de nodo"),
    ("Foros- Base Completa", "Foros", "Participaciones actor × evento con cargo, fecha y fuente", "Afiliación actor–evento"),
    ("Foros - Índice de Eventos", "Foros", "Catálogo de eventos: organizador, tipo, alcance", "Atributos de evento"),
    ("Foros - Nodos Centrales", "Foros", "Actores con participación recurrente en eventos", "Centralidad descriptiva"),
    ("Cruces Multiafiliacion", "Multiafiliación", "Casos de afiliaciones simultáneas tipificados", "Afiliación actor–organización"),
    ("Multiafiliación -", "Multiafiliación", "Portafolio sectorial de cada actor multiafiliado", "Afiliación actor–organización"),
]


AUTOR = "Darío Briceño"


def build_xlsx(counts):
    wb = openpyxl.load_workbook(RAW)  # modo escritura: preserva el libro completo
    wb.properties.creator = AUTOR
    wb.properties.lastModifiedBy = AUTOR
    ws = wb["Foros - Índice de Eventos"]
    fixed = 0
    for row in ws.iter_rows():
        evento = fmt(row[0].value)
        if evento in DATE_FIXES:
            row[2].value = DATE_FIXES[evento]
            row[2].number_format = "@"
            fixed += 1
    assert fixed == 2, f"se esperaban 2 correcciones de fecha, se hicieron {fixed}"

    resumen = wb.create_sheet("RESUMEN", 0)
    bold = openpyxl.styles.Font(bold=True)
    resumen.append(["RESUMEN — BBDD Actores de la Gobernanza del Litio en Chile"])
    resumen["A1"].font = bold
    resumen.append([])
    resumen.append(["Pestaña", "Tipo", "N entradas*", "Contenido", "Aporte a una red"])
    for c in resumen[3]:
        c.font = bold
    for name, tipo, desc, red in RESUMEN_SPEC:
        resumen.append([name, tipo, counts.get(name, ""), desc, red])
    resumen.append([])
    resumen.append(["* Entradas con contenido, excluyendo filas decorativas (banners ► y"])
    resumen.append(["  subtítulos). En 'Multiafiliación -' se cuentan afiliaciones actor–organización."])
    resumen.append(["  Fechas corregidas en 'Foros - Índice de Eventos': ver data_dictionary.md."])
    for col, width in zip("ABCDE", [30, 14, 12, 55, 28]):
        resumen.column_dimensions[col].width = width

    wb.save(OUT_XLSX)
    print(f"  {OUT_XLSX.relative_to(ROOT)}  (pestaña RESUMEN + {fixed} fechas corregidas)")


if __name__ == "__main__":
    print("Leyendo", RAW.relative_to(ROOT))
    wb = openpyxl.load_workbook(RAW, read_only=True, data_only=True)
    print("CSVs tidy:")
    counts = build_csvs(wb)
    wb.close()
    print("Excel procesado:")
    build_xlsx(counts)
    print("Listo.")
